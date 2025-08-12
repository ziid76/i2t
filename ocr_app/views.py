from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .forms import ImageUploadForm
from .models import OCRResult
from .utils import upload_to_s3, call_naver_ocr_api, save_ocr_result_to_file
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import io

def index(request):
    if request.method == 'POST':
        form = ImageUploadForm(request.POST, request.FILES)
        if form.is_valid():
            up = request.FILES.get('image_file')
            if not up:
                messages.error(request, '파일이 없습니다.')
                return redirect('index')

            # 업로드 스트림 → 메모리 복사(한 번만 read)
            blob = up.read()
            buf = BytesIO(blob)

            # S3 업로드 (file-like 사용)
            buf.seek(0)
            s3_url = upload_to_s3(buf, up.name)
            if not s3_url:
                messages.error(request, 'S3 업로드에 실패했습니다.')
                return redirect('index')

            # 모델 저장 (image_file은 비워둠)
            ocr_result = form.save(commit=False)
            ocr_result.image_file = None
            ocr_result.s3_url = s3_url

            # OCR 호출
            ocr_response = call_naver_ocr_api(s3_url)
            if not ocr_response:
                messages.error(request, 'OCR API 호출에 실패했습니다.')
                return redirect('index')

            ocr_result.ocr_result = ocr_response
            ocr_result.save()

            save_ocr_result_to_file(ocr_response)
            messages.success(request, 'OCR 처리가 완료되었습니다.')
            return redirect('ocr_result', pk=ocr_result.pk)
    else:
        form = ImageUploadForm()
    return render(request, 'ocr_app/index.html', {'form': form})

def ocr_result(request, pk):
    """OCR 결과 페이지"""
    try:
        ocr_result = OCRResult.objects.get(pk=pk)
        table_data = ocr_result.get_table_data()
        bounding_boxes = ocr_result.get_bounding_boxes()
        
        context = {
            'ocr_result': ocr_result,
            'table_data': table_data,
            'bounding_boxes': json.dumps(bounding_boxes),
            'image_url': ocr_result.image_file.url if ocr_result.image_file else ocr_result.s3_url
        }
        
        return render(request, 'ocr_app/result.html', context)
    
    except OCRResult.DoesNotExist:
        messages.error(request, '결과를 찾을 수 없습니다.')
        return redirect('index')

def get_ocr_results(request):
    """OCR 결과 목록 API"""
    results = OCRResult.objects.all().order_by('-created_at')[:10]
    data = []
    
    for result in results:
        data.append({
            'id': result.id,
            'created_at': result.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'image_url': result.image_file.url if result.image_file else result.s3_url,
            'has_result': bool(result.ocr_result)
        })
    
    return JsonResponse({'results': data})


def download_excel(request, pk):
    """OCR 결과를 엑셀 파일로 다운로드"""
    try:
        ocr_result = OCRResult.objects.get(pk=pk)
        table_data = ocr_result.get_table_data()
        
        if not table_data:
            messages.error(request, '다운로드할 테이블 데이터가 없습니다.')
            return redirect('ocr_result', pk=pk)
        
        # 엑셀 워크북 생성
        wb = Workbook()
        
        # 기본 시트 제거
        wb.remove(wb.active)
        
        # 각 테이블을 별도 시트로 생성
        for i, table in enumerate(table_data):
            sheet_name = f'Table_{i+1}'
            ws = wb.create_sheet(title=sheet_name)
            
            # 스타일 정의
            header_font = Font(bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # 테이블 데이터 입력
            for row_idx, row in enumerate(table, 1):
                for col_idx, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # 첫 번째 행은 헤더로 처리
                    if row_idx == 1:
                        cell.font = header_font
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 메모리에서 엑셀 파일 생성
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # HTTP 응답 생성
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ocr_table_result_{pk}.xlsx"'
        
        return response
        
    except OCRResult.DoesNotExist:
        messages.error(request, '결과를 찾을 수 없습니다.')
        return redirect('index')


